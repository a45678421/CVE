import openpyxl
import glob
import random
import os
import logging

# 確保目錄存在
log_dir = '../loading'
if not os.path.exists(log_dir):
    os.makedirs(log_dir)

# 設置日誌文件路徑
log_file_path = os.path.join(log_dir, 'log.txt')

# 配置日誌記錄器
logging.basicConfig(
    level=logging.INFO,
    filename=log_file_path,
    filemode='w',
    format='[%(levelname)1.1s %(asctime)s %(module)s:%(lineno)d] %(message)s',
    datefmt='%Y%m%d %H:%M:%S',
)

# 設置控制台處理器以顯示在終端
console = logging.StreamHandler()
console.setLevel(logging.INFO)
formatter = logging.Formatter('[%(levelname)1.1s %(asctime)s %(module)s:%(lineno)d] %(message)s', datefmt='%Y%m%d %H:%M:%S')
console.setFormatter(formatter)
logging.getLogger('').addHandler(console)

logging.info('Processing Excel files...')

# 查找上一層目錄中的 .xlsx 文件
xlsx_files = glob.glob('../Put_the_excel_file_here/*.xlsx')

if not xlsx_files:
    logging.error("No .xlsx files found in the specified directory.")
    raise FileNotFoundError("No .xlsx files found in the specified directory.")

# 打開找到的 .xlsx 文件
file_path = xlsx_files[0]  # 假設只處理找到的第一個文件
workbook = openpyxl.load_workbook(file_path)

# 選擇要填充的工作表
sheet = workbook.active

# 找到 Column1.redmine.upload 欄位的列索引，如果不存在則創建該欄位
header = [cell.value for cell in sheet[1]]
if 'Column1.redmine.upload' not in header:
    column_to_fill_index = len(header) + 1
    sheet.cell(row=1, column=column_to_fill_index, value='Column1.redmine.upload')
    logging.info('Column "Column1.redmine.upload" created.')
else:
    column_to_fill_index = header.index('Column1.redmine.upload') + 1

# 隨機填充從第二行開始的所有單元格，假設第一行是標題
for row in range(2, sheet.max_row + 1):
    sheet.cell(row=row, column=column_to_fill_index, value=random.choice([True, False]))

# 保存更改
workbook.save(file_path)

logging.info(f'已隨機填充文件 {file_path} 中的 Column1.redmine.upload 欄位。')
print(f'已隨機填充文件 {file_path} 中的 Column1.redmine.upload 欄位。')
